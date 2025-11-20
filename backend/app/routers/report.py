from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import date
from io import BytesIO
from sqlalchemy.orm import Session
import os

from app.db.database import get_db
from app.dependencies import get_current_user
from app.models.staffs import Staffs

router = APIRouter()


def get_reiwa_year(today: date) -> int:
    """西暦から令和年を計算する"""
    # 令和は2019年5月1日から開始
    reiwa_start = date(2019, 5, 1)
    if today < reiwa_start:
        # 令和開始前の場合は0を返す（実際には使用されない想定）
        return 0
    return today.year - 2018


def format_reiwa_date(today: date) -> str:
    """今日の日付を「令和　　年　　月　　日」形式に変換する"""
    reiwa_year = get_reiwa_year(today)
    # 「令和　　年　　月　　日」形式（数字の前に全角スペース）
    return f"令和　{reiwa_year}年　{today.month}月　{today.day}日"


def format_reiwa_year_month(year: int, month: int) -> str:
    """年月を「令和　　年　　月」形式に変換する"""
    target_date = date(year, month, 1)
    reiwa_year = get_reiwa_year(target_date)
    # 「令和　　年　　月」形式（数字の前に全角スペース）
    return f"令和　{reiwa_year}年　{month}月"


# 報告書テンプレートダウンロード
@router.get("/download-report-template")
def download_report_template(
    year: int = None,
    month: int = None,
    dayDirectCare: float = None,
    dayIndirectWork: float = None,
    dayBufferTime: float = None,
    dayOther: float = None,
    nightDirectCare: float = None,
    nightIndirectWork: float = None,
    nightBufferTime: float = None,
    nightOther: float = None,
    dayStaffCount: int | None = None,
    nightStaffCount: int | None = None,
    dayTotalHours: float | None = None,
    nightTotalHours: float | None = None,
    current_user: Staffs = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        # テンプレートファイルのパスを取得
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        template_path = os.path.join(base_dir, "templates", "report_template.xlsx")
        
        if not os.path.exists(template_path):
            raise HTTPException(status_code=404, detail="テンプレートファイルが見つかりません")
        
        # Excelファイルを読み込む
        wb = load_workbook(template_path)
        ws = wb.active  # アクティブなシートを取得
        
        # 今日の日付を取得
        today = date.today()
        formatted_date = format_reiwa_date(today)
        
        # AA2セルに日付を書き込む
        ws["AA2"] = formatted_date
        ws["AA2"].alignment = Alignment(
            horizontal="right",
            vertical="center",
        )
        
        # 事業所情報を取得
        office = current_user.office
        if office:
            # G6セルに事業所番号を書き込む
            ws["G6"] = office.jigyousyo_no or ""
            ws["G6"].alignment = Alignment(
                horizontal="left",
                vertical="center",
            )
            
            # G7セルに事業所名を書き込む
            ws["G7"] = office.name or ""
            ws["G7"].alignment = Alignment(
                horizontal="left",
                vertical="center",
            )
        
        # 選択されている年月をG53セルに書き込む
        if year and month:
            formatted_year_month = format_reiwa_year_month(year, month)
            ws["G53"] = formatted_year_month
            ws["G53"].alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
        
        # 選択した5日分の日中業務タイプ別割合を書き込む（％なし）
        # 日中: G57=直接介護, L57=間接業務, Q57=余裕時間, V57=その他
        if dayDirectCare is not None:
            ws["G57"] = dayDirectCare
            ws["G57"].alignment = Alignment(
                horizontal="right",
                vertical="center",
            )
        if dayIndirectWork is not None:
            ws["L57"] = dayIndirectWork
            ws["L57"].alignment = Alignment(
                horizontal="right",
                vertical="center",
            )
        if dayBufferTime is not None:
            ws["Q57"] = dayBufferTime
            ws["Q57"].alignment = Alignment(
                horizontal="right",
                vertical="center",
            )
        if dayOther is not None:
            ws["V57"] = dayOther
            ws["V57"].alignment = Alignment(
                horizontal="right",
                vertical="center",
            )
        
        # 選択した5日分の夜間業務タイプ別割合を書き込む（％なし）
        # 夜間: G64=直接介護, L64=間接業務, Q64=余裕時間, V64=その他
        if nightDirectCare is not None:
            ws["G64"] = nightDirectCare
            ws["G64"].alignment = Alignment(
                horizontal="right",
                vertical="center",
            )
        if nightIndirectWork is not None:
            ws["L64"] = nightIndirectWork
            ws["L64"].alignment = Alignment(
                horizontal="right",
                vertical="center",
            )
        if nightBufferTime is not None:
            ws["Q64"] = nightBufferTime
            ws["Q64"].alignment = Alignment(
                horizontal="right",
                vertical="center",
            )
        if nightOther is not None:
            ws["V64"] = nightOther
            ws["V64"].alignment = Alignment(
                horizontal="right",
                vertical="center",
            )

        # 調査対象人数をC55・C62に書き込む
        if dayStaffCount is not None:
            ws["C55"] = f"① 日中\u3000調査対象人数\u3000{dayStaffCount}人"
            ws["C55"].alignment = Alignment(
                horizontal="left",
                vertical="center",
            )
        if nightStaffCount is not None:
            ws["C62"] = f"② 夜間\u3000調査対象人数\u3000{nightStaffCount}人"
            ws["C62"].alignment = Alignment(
                horizontal="left",
                vertical="center",
            )

        # 合計計測時間をL60・L66に書き込む
        if dayTotalHours is not None:
            ws["L60"] = dayTotalHours
            ws["L60"].alignment = Alignment(
                horizontal="right",
                vertical="center",
            )
        if nightTotalHours is not None:
            ws["L66"] = nightTotalHours
            ws["L66"].alignment = Alignment(
                horizontal="right",
                vertical="center",
            )
        
        # 一時的にメモリに格納
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)  # 読みだし位置を先頭に戻す
        
        # 効率よくダウンロードするためのレスポンス
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=report_template.xlsx"},
        )
    except HTTPException:
        raise
    except Exception as e:
        print("❌ テンプレートダウンロードエラー:", str(e))
        raise HTTPException(status_code=500, detail="テンプレートのダウンロードに失敗しました")

