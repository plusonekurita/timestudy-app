from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from pydantic import BaseModel
from datetime import date, datetime
from typing import List
from app.db.database import get_db
from app.models.time_record import TimeRecord
from app.utils.time_conversion import convert_with_time_split
from app.services.excel_writer import load_excel_template, write_records_to_excel
from fastapi.responses import FileResponse
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl.styles import Alignment

router = APIRouter()

class RecordEntry(BaseModel):
    id: str
    no: int
    type: str
    name: str
    label: str
    startTime: str
    endTime: str
    duration: int

class OfficeEntry(BaseModel):
    id: int
    name: str

class StaffEntry(BaseModel):
    id: int
    staffCode: str
    office: OfficeEntry

class TimeRecordRequest(BaseModel):
    staff: StaffEntry
    record_date: date
    record: List[RecordEntry]

class TimeRecordRangeRequest(BaseModel):
    staff_id: int
    start_date: date # YYYY-MM-DD
    end_date: date # YYYY-MM-DD

@router.post("/save-time-records")
def create_time_record(request: TimeRecordRequest, db: Session = Depends(get_db)):
    try:
        # 既に同じ staff_id + record_date のレコードがあるか確認
        existing = db.query(TimeRecord).filter_by(
            staff_id=request.staff.id,
            record_date=request.record_date
        ).first()

        if existing:
            # 既存レコードがある場合 → 上書き更新
            existing.record = [r.model_dump() for r in request.record] # レコードの上書き
            db.commit()
            db.refresh(existing)
            return {"message": "Time record updated", "id": existing.id}
        else:
            # 新規レコードを作成
            new_record = TimeRecord(
                staff_id=request.staff.id,
                record_date=request.record_date,
                record=[r.model_dump() for r in request.record]
            )
            db.add(new_record)
            db.commit()
            db.refresh(new_record)
            return {"message": "Time record created", "id": new_record.id}

    except Exception as e:
        print("エラー:", str(e))
        raise HTTPException(status_code=500, detail="DB保存に失敗しました")

@router.post("/get-time-records")
def get_range_record(request: TimeRecordRangeRequest, db: Session = Depends(get_db)):
    print(request)
    try:
        query = db.query(TimeRecord).filter(
            TimeRecord.staff_id == request.staff_id
        )

        # end_date がない or start_date と同じ → 単一日付の扱い
        if not request.end_date or request.start_date == request.end_date:
            query = query.filter(TimeRecord.record_date == request.start_date)
            records = query.first()

            if not records:
                return {
                    "message": "記録は見つかりませんでした。",
                    "records": None
                }

            return {
                "message": "1件の記録を取得しました。",
                "records": [records.__dict__]
            }

        # 範囲指定 → 複数件
        else:
            query = query.filter(
                TimeRecord.record_date >= request.start_date,
                TimeRecord.record_date <= request.end_date
            )
            records = query.all()

            if not records:
                return {
                    "message": "記録は見つかりませんでした。",
                    "records": []
                }

            return {
                "message": f"{len(records)} 件の記録を取得しました。",
                "records": [r.__dict__ for r in records]
            }

    except Exception as e:
        print("❌ 記録取得エラー:", str(e))
        raise HTTPException(status_code=500, detail="記録取得中にエラーが発生しました")


@router.post("/convert-time-records")
def convert_time_records(request: TimeRecordRequest):
    try:
        # Pydantic RecordEntry → dict に変換
        records_as_dict = [r.model_dump() for r in request.record]

        # 時間変換処理
        converted = convert_with_time_split(records_as_dict)

        return {
            "message": "変換に成功しました。",
            "converted": converted
        }
    except Exception as e:
        print("❌ 変換エラー:", str(e))
        raise HTTPException(status_code=500, detail="変換処理に失敗しました")


# エクセルファイル出力
@router.post("/export_excel")
def export_excel(request: TimeRecordRequest):

    try:
        wb = load_excel_template()
        ws = wb["sheet1"]

        # 整形済みデータと開始時間を取得
        records_as_dict = [r.model_dump() for r in request.record]
        converted = convert_with_time_split(records_as_dict)
        start_timestr = list(converted.keys())[0]
        start_hour = int(start_timestr.split(":")[0])

        # 〇タイムスタディ記録の書き込み
        write_records_to_excel(ws, converted, start_hour)

        # 〇施設名
        ws["H2"] = request.staff.office.name
        ws["H2"].alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

        # 職員ID
        ws["H3"] = request.staff.staffCode
        ws["H3"].alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

        # 〇調査実施日--
        record_date = request.record_date

        # 日本語の曜日表記に変換
        weekday_jp = ["月", "火", "水", "木", "金", "土", "日"][record_date.weekday()]

        # "8月1日（金）" のような書式を作成
        formatted_date = f"{record_date.month}月    {record_date.day}日 （   {weekday_jp}曜日）"

        # Excelの4Hセルに書き込み
        ws["H4"] = formatted_date
        ws["H4"].alignment = Alignment(
            horizontal="left",   # 中央揃え（横）
            vertical="center",     # 中央揃え（縦）
            wrap_text=False,       # 折り返しなし
            shrink_to_fit=True     # セル幅に合わせて自動縮小
        )


        # 〇実勤務時間
        # 最小の startTime と最大の endTime を求める
        start_times = [datetime.fromisoformat(r.startTime) for r in request.record]
        end_times = [datetime.fromisoformat(r.endTime) for r in request.record]

        if start_times and end_times:
            min_start = min(start_times)
            max_end = max(end_times)

            # "HH:MM~HH:MM" の形式に整形
            start_str = min_start.strftime("%H：%M")
            end_str = max_end.strftime("%H：%M")
            time_range = f"{start_str}        ～        {end_str}"

            # S4 に書き込む（中央寄せ）
            ws["S4"] = time_range
            ws["S4"].alignment = Alignment(horizontal="center", vertical="center")

        # 一時的にメモリに格納
        buffer = BytesIO()
        wb.save(buffer)
        print("保存後のバッファサイズ:", buffer.tell())
        buffer.seek(0) # 読みだし位置を先頭に戻す

        # 効率よくダウンロードするためのレスポンス
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=timeStudy_sheet.xlsx"},
        )

    except Exception as e:
            print("❌ エクスポートエラー:", str(e))
            raise HTTPException(status_code=500, detail="Excel出力に失敗しました")