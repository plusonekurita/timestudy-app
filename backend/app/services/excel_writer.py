from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import os

def load_excel_template():
    # このファイルから見たテンプレートの相対パスを解決
    base_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_dir, "..", "templates", "timeStudy_template.xlsx")

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"テンプレートファイルが見つかりません: {template_path}")

    # Excelファイルを読み込む
    wb = load_workbook(template_path)
    return wb



def write_records_to_excel(ws: Worksheet, converted: dict, start_hour: int):
    time_header_cols = ["F", "L", "R", "X", "AD", "AJ"]

    # 列マッピング（0〜35）
    base_col_indices = {
        0: "F", 1: "G", 2: "H", 3: "I", 4: "J", 5: "K",
        6: "L", 7: "M", 8: "N", 9: "O", 10: "P", 11: "Q",
        12: "R", 13: "S", 14: "T", 15: "U", 16: "V", 17: "W",
        18: "X", 19: "Y", 20: "Z", 21: "AA", 22: "AB", 23: "AC",
        24: "AD", 25: "AE", 26: "AF", 27: "AG", 28: "AH", 29: "AI",
        30: "AJ", 31: "AK", 32: "AL", 33: "AM", 34: "AN", 35: "AO",
    }

    # ヘッダー書き込み済みトラッカー
    page1_written = set()
    page2_written = set()

    for timestr, entries in converted.items():
        hour, minute = map(int, timestr.split(":"))
        hour_offset = hour - start_hour
        if hour_offset < 0:
            hour_offset += 24

        minute_block = minute // 10
        time_index = hour_offset * 6 + minute_block

        page_num = time_index // 36  # 0: 1ページ目, 1: 2ページ目
        local_index = time_index % 36
        col = base_col_indices.get(local_index)
        if not col:
            continue

        # ヘッダーに「○時台」を書き込む
        group_index = local_index // 6
        header_col = time_header_cols[group_index]
        if page_num == 0:
            if header_col not in page1_written:
                ws[f"{header_col}9"] = f"{start_hour + group_index}時台"
                page1_written.add(header_col)
        else:
            if header_col not in page2_written:
                ws[f"{header_col}43"] = f"{start_hour + 6 + group_index}時台"
                page2_written.add(header_col)

        # データ書き込み
        for worker_str, value in entries.items():
            worker_num = int(worker_str)
            row = 10 + worker_num if page_num == 0 else 44 + worker_num

            if value != 0:
                ws[f"{col}{row}"] = value
